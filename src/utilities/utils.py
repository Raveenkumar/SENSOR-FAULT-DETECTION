from email.mime.application import MIMEApplication
from box import ConfigBox
from pathlib import Path
import json
import zipfile
import shutil
import hashlib
import pickle
import dill
import os,sys
from typing import Dict, Any
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
import xgboost as xgb
from sklearn.metrics import confusion_matrix,ConfusionMatrixDisplay
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
from sklearn.metrics import accuracy_score, recall_score, precision_score, f1_score, roc_auc_score,classification_report
from sklearn.model_selection import RandomizedSearchCV
from sklearn.base import BaseEstimator
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from dotenv import load_dotenv
from src.logger import logger
from src.exception import SensorFaultException
from src.entity.config_entity import (BaseArtifactConfig,
                                      ModelTrainerConfig,
                                      ModelTunerConfig,
                                      PredictionRawDataValidationConfig,
                                      ModelEvaluationConfig,
                                      PredictionPipelineConfig)
from src.entity.artifact_entity import ModelTunerArtifacts

matplotlib.use('Agg')  # Switch to Agg backend for non-interactive use

load_dotenv()

def create_folder_using_folder_path(folder_path:Path):
    """create_folder_using_folder_path :Used for create a folder using folder path

    Args:
        folder_path (Path): folder_path to create

    Raises:
        error_message: CustomException
    """
    try:
        os.makedirs(folder_path,exist_ok=True)
        logger.info(f"create folder using folder path :: Status:Successful :: folder_path:{folder_path}")
    except Exception as e:
        error_message =  SensorFaultException(error_message=str(e),error_detail=sys)
        logger.error(f"create folder using folder path :: Status:Failed :: folder_path:{folder_path} :: Error:{error_message}")
        raise error_message

def create_folder_using_file_path(file_path:Path):
    """create_folder_using_file_path :Used for create a folder using file path

    Args:
        file_path (Path): file path we  can get folder name

    Raises:
        error_message: CustomException
    """
    try:
        folder_name,_ = os.path.split(file_path)
        os.makedirs(folder_name,exist_ok=True)
        logger.info(f"create folder using file path :: Status:Successful :: folder_path:{file_path}")
    except Exception as e:
        error_message =  SensorFaultException(error_message=str(e),error_detail=sys)
        logger.error(f"create folder using file path :: Status:Failed :: folder_path:{file_path} :: Error:{error_message}")
        raise error_message

def read_json(file_path:Path) -> ConfigBox:
    """read_json: used for reading json file 

    Args:
        file_path (Path): file_path

    Returns:
        ConfigBox: Provide ConfigBox (contains json data)
    """
    try:
        #check file path exist or not
        if os.path.exists(file_path):
            with open(file_path,mode='r') as file:
                json_data = json.load(file)
            logger.info(f"read json data :: file_path:{file_path} :: Status:Successful")    
        return ConfigBox(json_data)    
    
    except Exception as e:
        error_message =  SensorFaultException(error_message=str(e),error_detail=sys)
        logger.error(f"read json data :: file_path:{file_path} :: Status:Failed :: Error:{error_message}")
        raise error_message  

def save_json(file_path:Path, file_obj:object):
    try:
        #check file path exist or not
        with open(file_path, 'w') as json_file:
            json.dump(file_obj, json_file, indent=4)
        logger.info(f"save dict data into json :: file_path:{file_path} :: Status:Successful")    
       
    except Exception as e:
        error_message =  SensorFaultException(error_message=str(e),error_detail=sys)
        logger.error(f" save dict data into json :: file_path:{file_path} :: Status:Failed :: Error:{error_message}")
        raise error_message     

def read_csv_file(file_path:Path) -> pd.DataFrame:
    """read_csv_file :: Used for read the csv file using pandas

    Args:
        file_path (Path): File path of the file

    Raises:
        SensorFaultException: Custom Exception

    Returns:
        pd.DataFrame: dataframe
    """
    try:
        #check file path exist or not
        if os.path.exists(file_path):
            dataframe = pd.read_csv(file_path)
            logger.info(f"read csv file : file_path: {file_path} : Status: Successful")    
        return dataframe    
    
    except Exception as e:
        error_message =  SensorFaultException(error_message=str(e),error_detail=sys)
        logger.error(f"read csv data :: file_path:{file_path} :: Status:Failed :: Error:{error_message}")
        os._exit(1)
        raise error_message
        
def style_excel(excel_filename):
    """style_excel : Used for style the append_log_to_excel method

    Args:
        excel_filename (_type_): Excel file name
    """
    
    # Load the workbook and select the active worksheet
    wb = load_workbook(excel_filename)
    ws = wb.active

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Define header fill color and font style
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow color
    header_font = Font(bold=True, color='000000')  # Black color for text

    # Apply styles to the header row
    for cell in ws[1]: # type: ignore
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Apply border to all data cells and adjust column widths
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column, max_row=ws.max_row):# type: ignore
        for cell in row:
            cell.border = thin_border

    # Adjust column widths based on the maximum length of the data in each column
    for column in ws.columns:# type: ignore
        max_length = 0
        column_letter = column[0].column_letter  # Get the column letter (e.g., 'A', 'B')
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Add some padding
        ws.column_dimensions[column_letter].width = adjusted_width# type: ignore

    # Save the styled workbook
    wb.save(excel_filename)

def style_excel_model_result(excel_filename: str) -> None:
    """style_excel : Used for styling the Excel file and adding feedback dropdown.

    Args:
        excel_filename (str): Path to the Excel file
    """
    wb = load_workbook(excel_filename)
    ws = wb.active

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Define header fill color and font style
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow color
    header_font = Font(bold=True, color='000000')  # Black color for text

    # Apply styles to the header row
    for cell in ws[1]:  # type: ignore
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Apply border to all data cells and adjust column widths
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column, max_row=ws.max_row):  # type: ignore
        for cell in row:
            cell.border = thin_border

    # Adjust column widths based on the maximum length of the data in each column
    for column in ws.columns:  # type: ignore
        max_length = 0
        column_letter = column[0].column_letter  # Get the column letter (e.g., 'A', 'B')
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Add some padding
        ws.column_dimensions[column_letter].width = adjusted_width  # type: ignore

    # Add DataValidation for the "Feedback" column (Assumed to be column D in this case)
    feedback_validation = DataValidation(
        type="list", 
        formula1='"Working,NotWorking"', 
        allow_blank=True
    )
    feedback_validation.error = 'Invalid feedback'
    feedback_validation.errorTitle = 'Invalid Input'
    feedback_validation.prompt = 'Select from the list'
    feedback_validation.promptTitle = 'Feedback'

    # Assuming Feedback is the fourth column (D), apply dropdown to all rows in column D
    ws.add_data_validation(feedback_validation) # type: ignore
    feedback_validation.add(f'D2:D{ws.max_row}') # type: ignore

    # Save the styled and updated workbook
    wb.save(excel_filename)

def append_log_to_excel(filename:str, status:str, status_reason: str, remark:str, excel_filename:Path):
    """append_log_to_excel :Used for store logs of validating data

    Args:
        filename (str): validating file name
        status (str): validation status
        remark (str): any remark related to validation
        excel_filename (str, optional): validation store file name
    """
    # Get the current date for the DATE column
    current_date = datetime.now().strftime('%Y-%m-%d')
    try:
        # Load the existing data into a DataFrame if the file exists
        existing_df = pd.read_excel(excel_filename)
        
        # Define slno automatically based on the number of existing rows
        slno = len(existing_df) + 1
                
        # Create a DataFrame for the new log entry
        new_data = {
            'SLNO': [slno],
            'DATE': [current_date],
            'FILENAME': [filename],
            'STATUS': [status],
            'STATUS_REASON': [status_reason],
            'REMARK': [remark]
        }
        new_df = pd.DataFrame(new_data)
        
        # Concatenate the existing DataFrame with the new DataFrame
        updated_df = pd.concat([existing_df, new_df], ignore_index=True)
        
        # Save the updated DataFrame to the Excel file
        with pd.ExcelWriter(excel_filename, engine='openpyxl', mode='w') as writer:
            updated_df.to_excel(writer, index=False)
        
        # Style the Excel file
        style_excel(excel_filename)
        # logger.info(f"File:log_file.xlsx Data added :: Status:Successful :: Data:{new_df.to_dict()}")
        logger.info(f"append_log_to_excel :: file_name:{excel_filename} :: Status:Data added")
    except FileNotFoundError:
        # If the file does not exist, create a new Excel file with slno starting from 1
        new_data = {
            'SLNO': [1],
            'DATE': [current_date],
            'FILENAME': [filename],
            'STATUS': [status],
            'STATUS_REASON': [status_reason],
            'REMARK': [remark]
        }
        new_df = pd.DataFrame(new_data)
        new_df.to_excel(excel_filename, index=False, engine='openpyxl')
        
        # Style the Excel file
        style_excel(excel_filename)
        logger.info(f"append_log_to_excel :: file_name:{excel_filename} :: Status:Created")
        # logger.info(f"File:log_file.xlsx Data added :: Status:Successful :: Data:{new_df.to_dict()}")
        logger.info(f"append_log_to_excel :: file_name:{excel_filename} :: Status:Data added")

def format_as_s3_path(path:Path) -> str:
    """format_as_s3_path :Used for convert path to s3_path format

    Args:
        path (Path): it can be window path or other path format

    Raises:
        error_message: Custom Exception

    Returns:
        str: s3_path
    """
    try:
        # if path is file path
        if os.path.isfile(path):
            s3_path  = str(path)
            s3_path = s3_path.replace('\\','/')
        else:
            s3_path  = str(path)        
            s3_path = s3_path.replace('\\','/')+'/'
        logger.info(f'convert path:{path} to s3_path:{s3_path} successfully')
        return s3_path
    except Exception as e:
        error_message =  SensorFaultException(error_message=str(e),error_detail=sys)
        logger.error(f"format_as_s3_path convert path to s3_path :: file_path:{path} :: Status:Failed :: Error:{error_message}")
        raise error_message  
    
def delete_artifact_folder() -> None:
    """delete_artifact_folder :Used for drop the folder artifact after completed prediction or training process is completed

    Raises:
        error_message: Custom Exception
    """
    try:
        shutil.rmtree(BaseArtifactConfig.artifact_base_dir)
        logger.info("delete artifact folder:: Status: Successfully")
    except Exception as e:
        error_message =  SensorFaultException(error_message=str(e),error_detail=sys)
        logger.error(f"delete_artifact_folder  ::  Status:Failed :: Error:{error_message}")
        raise error_message

def save_obj(file_path:Path,obj: object):
    """save_obj Used: save the model to pkl objects

    Args:
        file_path (Path): file path for save the object
        obj (object): model object

    Raises:
        SensorFaultException: Custom Exception
    """
    try:
        dir_name = os.path.dirname(file_path)
        os.makedirs(dir_name,exist_ok=True)
        if os.path.basename(file_path).endswith(".pkl"):
            with open(file_path,'wb') as file_obj:
                pickle.dump(obj,file=file_obj) 
        else:
            with open(file_path,'wb') as file_obj:
                dill.dump(obj,file=file_obj)        
        logger.info(msg=f'object saved :: Status: Success :: path: {file_path}')    
        
    except Exception as e:
        error_message =  SensorFaultException(error_message=str(e),error_detail=sys)
        logger.error(f"save_obj  ::  Status:Failed :: Error:{error_message}")
        raise error_message
    
def load_obj(file_path:Path)-> object:
    """load_obj :Used for load pkl object

    Args:
        file_path (Path): file path of object

    Raises:
        SensorFaultException: Custom Exception

    Returns:
        object: pkl object
    """
    try:
        if os.path.exists(file_path):
            if os.path.basename(file_path).endswith(".pkl"):
                with open(file_path,mode='rb') as file_obj:
                    result = pickle.load(file=file_obj)
            elif os.path.basename(file_path).endswith(".xgb"):
                loaded_model = xgb.Booster()
                loaded_model.load_model(file_path)
                result = loaded_model
            else:
                with open(file_path,mode='rb') as file_obj:
                    result = dill.load(file=file_obj)        
            logger.info(msg=f'object loaded :: Status: Success :: path: {file_path}') 
               
            return result     
        else:
            e = Exception(f'object file path not exist path: {file_path}')    
            logger.info(msg=e)
            raise e
    
    except Exception as e:
        error_message =  SensorFaultException(error_message=str(e),error_detail=sys)
        logger.error(f"load object   ::  Status:Failed :: Error:{error_message}")
        raise error_message    
     
def model_result(model_name:str,model:RandomizedSearchCV,X_train:pd.DataFrame,X_test:pd.DataFrame,y_train:pd.DataFrame,y_test:pd.DataFrame)-> tuple[dict[str, Any],BaseEstimator,dict[str, Any]]:
    """model_result :Used for store model result data in dict format

    Args:
        model (RandomizedSearchCV): RandomizedSearch CV object
        X_train (DataFrame): X train
        X_test (DataFrame): X_test
        y_train (DataFrame): y_train
        y_test (DataFrame): y_test

    Raises:
        error_message: Custom Exception

    Returns:
       tuple[dict[str, Any],RandomizedSearchCV,dict[str, Any]]: mlflow_dict, model,results
    """
    try: 
        y_pred = model.predict(X_test)  
        
        # get classification report for individuals like recall of 0 ,1
        classification_report_ :Dict[str, Dict[str, Any]]=  classification_report( y_test,y_pred,output_dict=True) # type: ignore
        # save the confusion matrix image
        create_folder_using_folder_path(ModelEvaluationConfig.confusion_matrixes_path)
        confusion_matrix_image_path = ModelEvaluationConfig.confusion_matrixes_path / f"{model_name}.png"
        logger.info(f"confusion matrix image path: {confusion_matrix_image_path}")
        save_confusion_matrix_image(y_test=y_test,y_pred=y_pred,image_path=confusion_matrix_image_path)
        
        logger.info(f'Classification_report: {classification_report_}')
        mlflow_dict = {"model": model.best_estimator_,
                        "param" : model.best_params_,
                        'training_score': model.score(X_train, y_train),
                        'test_score': model.score(X_test, y_test), 
                        'auc_score': roc_auc_score(y_test, y_pred),
                        "overall_recall_score" : recall_score(y_test, y_pred, zero_division=0),
                        "recall_1": classification_report_['1']['recall'] ,
                        "recall_0": classification_report_['0']['recall'],
                        "overall_precision" : precision_score(y_test, y_pred, zero_division=0),
                        "precision_1": classification_report_['1']['precision'] ,
                        "precision_0": classification_report_['0']['precision'],
                        "confusion_matrix_image_path": confusion_matrix_image_path
                    }
        
        result_dict = {
            'best_param': model.best_params_,
            'training_score': model.score(X_train, y_train),
            'test_score': model.score(X_test, y_test), 
            'accuracy': accuracy_score(y_test, y_pred),
            'recall': recall_score(y_test, y_pred, zero_division=0),
            'precision': precision_score(y_test, y_pred, zero_division=0),
            'f1_score': f1_score(y_test, y_pred, zero_division=0),
            'auc_score': roc_auc_score(y_test, y_pred)
        }
        
        logger.info(f"model_result: {result_dict}")
        
        return mlflow_dict, model.best_estimator_, result_dict
    
    except Exception as e:
        error_message = SensorFaultException(error_message=str(e),error_detail=sys)
        logger.error(msg=f"model_result :: Status:Failed :: Error:{error_message}")
        raise error_message

def save_model_result_excel(df: pd.DataFrame, excel_file_path: Path) -> None:
    """save_model_result_excel :Used for store the model results in excel file 

    Args:
        df (pd.DataFrame): DataFrame containing the model results
        excel_file_path (str): Path where the excel file will be saved
    """
    try:
        # Save the DataFrame to Excel
        df.to_excel(excel_file_path, index=False, engine='openpyxl')

        # Style the Excel file and add dropdown
        style_excel(excel_file_path)
        logger.info("save_model_result_excel :: Status:created-result data added")

    except Exception as e:
        error_message = SensorFaultException(error_message=str(e), error_detail=sys)
        logger.error(f"save_model_result_excel :: Status:Failed :: Error:{error_message}")
        raise error_message

def save_model_result_feedback_excel(df: pd.DataFrame, excel_file_path: str) -> None:
    """save_model_result_excel :Used for store the model results in excel file 

    Args:
        df (pd.DataFrame): DataFrame containing the model results
        excel_file_path (str): Path where the excel file will be saved
    """
    try:
        # Add the "Feedback" column with default empty values
        df[PredictionPipelineConfig.feedback_column_name] = ""

        # Save the DataFrame to Excel
        df.to_excel(excel_file_path, index=False, engine='openpyxl',float_format="%.2f")

        # Style the Excel file and add dropdown
        style_excel_model_result(excel_file_path)
        logger.info("save_model_result_excel :: Status:created-result data added")

    except Exception as e:
        error_message = SensorFaultException(error_message=str(e), error_detail=sys)
        logger.error(f"save_model_result_excel :: Status:Failed :: Error:{error_message}")
        raise error_message

def save_models_data(all_model_objects_path:Path,
                best_model_object_path:Path,
                model_tuner_artifacts:ModelTunerArtifacts):
    """save_models_data :Used for save the model results & model objects

    Args:
        all_model_objects_path (Path): path of  files of store all models data
        best_model_object_path (Path): path of  files of store best model data
        model_tuner_artifacts (ModelTunerArtifacts): contains  all models results and best model results

    Raises:
        error_message: Custom Exception
    """
    try:
        logger.info(f'save_models_data:: create folder for all_model_objects_path,best_model_object_path if not exist:{all_model_objects_path,best_model_object_path}')
        create_folder_using_folder_path(all_model_objects_path)
        create_folder_using_folder_path(best_model_object_path)
                
        all_models_data = model_tuner_artifacts.all_models_data
        best_model_data = model_tuner_artifacts.best_model_data

        
        for model_name,model_obj in all_models_data[0].items():
            file_name = model_name+".pkl"
            
            all_model_file_path = all_model_objects_path / file_name
            
            # final_file_path = find_final_path(experiment_file_path=exp_all_model_file_path,stable_file_path=stable_all_model_file_path)
            
            final_file_path = all_model_file_path
            
            #save model object
            logger.info(f"save_models_data :: saved the model {file_name} into {all_model_objects_path}")
            save_obj(final_file_path,model_obj)
            
        final_path = os.path.dirname(final_file_path)
        json_file_path = Path(os.path.join(final_path,ModelTrainerConfig.all_model_result_json_file_name))
        logger.info(f"save_models_data :: saved the model json data {ModelTrainerConfig.all_model_result_json_file_name} into {final_path}") 
        save_json(json_file_path,all_models_data[1])   
            
        
        file_name = ModelTrainerConfig.best_model_name
        # exp_best_model_path = experiment_best_model_object_path / file_name
        # stable_best_file_path = stable_best_model_object_path / file_name
        
        best_model_obj_file_path = best_model_object_path / file_name
        # final_file_path = find_final_path(experiment_file_path=exp_best_model_path,stable_file_path=stable_best_file_path)
        
        final_file_path = best_model_obj_file_path
        
        #save model object
        best_model_obj = best_model_data[0]
        best_model_results = best_model_data[1]
        
        #save model object
        logger.info(f"save_models_data :: saved the best model {file_name} into {best_model_object_path}")
        save_obj(final_file_path,best_model_obj)
        final_path = os.path.dirname(final_file_path)
        json_file_path = Path(os.path.join(final_path,ModelTrainerConfig.best_model_json_file_name)) 
        logger.info(f"save_models_data :: saved the best model json data {ModelTrainerConfig.best_model_json_file_name} into {final_path}")
        save_json(json_file_path,best_model_results)   
 
    except Exception as e:
            error_message = SensorFaultException(error_message=str(e),error_detail=sys)
            logger.error(msg=f"save_model_result_excel :: Status:Failed :: Error:{error_message}")
            raise error_message 

def find_final_path(experiment_file_path:Path,stable_file_path:Path) -> Path:
    """find_final_path :Used find the folder path weather we store in experiments or stable's data

    Args:
        experiment_file_path (Path): Folder path
        stable_file_path (Path): Folder path

    Raises:
        error_message: Custom Exception

    Returns:
        Path: Final path
    """
    try:
        if not os.path.exists(stable_file_path):
            return stable_file_path 
        else:    
            return experiment_file_path
                    
    except Exception as e:
            error_message = SensorFaultException(error_message=str(e),error_detail=sys)
            logger.error(msg=f"save_object_to_experiment_and_stable :: Status:Failed :: Error:{error_message}")
            raise error_message
                                             
def proper_conversion_for_excel_file(result_dict:dict) -> pd.DataFrame:
    """proper_conversion_for_excel_file :Used for change data into proper for format for excel file

    Args:
        result_dict (dict): result dict

    Raises:
        error_message: Custom Exception

    Returns:
        pd.DataFrame: proper data in dataframe format
    """
    try:
            data = []
            for cluster_name, models in result_dict.items():
                for model_name, metrics in models.items():
                    # Add cluster name, model name, and metrics to the list
                    data.append({
                        'cluster': cluster_name,
                        'model_name': model_name,
                        **metrics  # Unpack the metrics dictionary into separate columns
                    })

            # Convert the collected data into a DataFrame
            df = pd.DataFrame(data) 
            logger.info(msg=f"proper_conversion_for_excel_file :: Status:Success :: result_dict: {result_dict}")
            
            return df
    except Exception as e:
            error_message = SensorFaultException(error_message=str(e),error_detail=sys)
            logger.error(msg=f"proper_conversion_for_excel_file :: Status:Failed :: Error:{error_message}")
            raise error_message                              
              
def get_dir_path_from_file_path(file_path:Path) -> Path:
    """get_dir_path_from_file_path Used getting the dir path from the file path

    Args:
        file_path (Path): file_path

    Raises:
        error_message: Custom Exception

    Returns:
        Path: dir path
    """
    try:
        dir_path = os.path.dirname(file_path)    
        logger.info(msg=f"get_dir_path_from_file_path :: Status:Success :: file_path:{file_path} :: folder_path:{dir_path} ")
        return Path(dir_path)
    except Exception as e:
            error_message = SensorFaultException(error_message=str(e),error_detail=sys)
            logger.error(msg=f"get_dir_path_from_file_path :: Status:Failed :: file_path:{file_path} :: Error:{error_message}")
            raise error_message                              
                
def copy_file(src_file_path:Path,dst_folder_path:Path):
    """copy_file :Used for copy file from source file path to destination folder path

    Args:
        src_file_path (Path): Path of source file path
        dst_folder_path (Path): destination folder path

    Raises:
        error_message: Custom Exception
    """
    try:
        shutil.copy2(src=src_file_path,dst=dst_folder_path)
        logger.info(msg=f"copy_file :: Status:Success :: src_file_path:{src_file_path} :: dst_folder_path:{dst_folder_path} ")
        
    except Exception as e:
            error_message = SensorFaultException(error_message=str(e),error_detail=sys)
            logger.error(msg=f"get_dir_path_from_file_path :: Status:Failed :: Error:{error_message}")
            raise error_message    

def create_zip_from_folder(folder_path:Path, output_zip_path:Path):
    """create_zip_from_folder :Used for create zip for folder data

    Args:
        folder_path (Path): folder path
        output_zip_path (Path): zip file path

    Raises:
        error_message: Custom Exception
    """
    try:
        # Create a zip file at the specified output path
        with zipfile.ZipFile(output_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Walk through the folder, including all subdirectories and files
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    # Add file to the zip, preserving the folder structure
                    arcname = os.path.relpath(file_path, folder_path)  # Relative path inside the zip
                    zipf.write(file_path, arcname)
                    logger.info(f'Added {file_path} as {arcname}')                    
        logger.info(f"create_zip_from_folder :: Status:Success :: folder_path__path:{folder_path} :: output_zip_path:{output_zip_path}")
    except Exception as e:
            error_message = SensorFaultException(error_message=str(e),error_detail=sys)
            logger.error(msg=f"create_zip_from_folder :: Status:Failed :: folder_path__path:{folder_path} :: output_zip_path:{output_zip_path} :: Error:{error_message}")
            raise error_message
                                
def models_auc_threshold_satisfied() -> bool:
    """ models_auc_threshold_satisfied:Used for check all three models auc score are satisfied the auc threshold value or not 

    Raises:
        error_message: Custom Exception

    Returns:
        bool: return True if satisfied else return False
    """
    try:
        i=0
        best_mode_results = read_json(ModelTrainerConfig.final_best_model_results_json_file_path)
        for cluster_name, model_data in best_mode_results.items():
            for model_name, model_results in model_data.items():
                if model_results['auc_score'] >= ModelTunerConfig.auc_score_threshold_value:
                    logger.info(f"{cluster_name}:{model_name}:{model_results['auc_score']}:: Status:Satisfied")
                    i+=1
                else:
                    logger.info(f"{cluster_name}:{model_name}:{model_results['auc_score']}:: Status:Not Satisfied")
        if i==3:
            return True
        else:
            return False                
    except Exception as e:
            error_message = SensorFaultException(error_message=str(e),error_detail=sys)
            logger.error(msg=f"models_auc_threshold_satisfied :: Status:Failed :: error_message:{error_message}")
            raise error_message     

def clear_artifact_folder(artifact_dir:str) -> None:
    """
    Clears the contents of the 'artifacts' directory located in the specified project root.

    This function checks if the 'artifacts' directory exists within the given project root path.
    If it does, it iterates through its contents and removes all files, symbolic links, and subdirectories.

    Args:
        project_root (str): The root directory of the project where the 'artifacts' folder is located.

    Raises:
        SensorFaultException: If an error occurs while attempting to delete files or directories, 
        an exception is raised with details about the failure.
    """
    # artifact_dir = os.path.join(project_root, 'artifacts')  # 'artifacts' folder in the root
    # logger.info(f'project root folder: {project_root}')
    logger.info(f'artifact_dir folder: {artifact_dir}')
    
    try:
        # Check if the directory exists
        if os.path.exists(artifact_dir):
            # Loop through the contents of the artifact directory
            for filename in os.listdir(artifact_dir):
                file_path = os.path.join(artifact_dir, filename)
                try:
                    if os.path.isfile(file_path) or os.path.islink(file_path):
                        os.unlink(file_path)  # Remove the file or symbolic link
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)  # Remove the directory and all its contents
                except Exception as e:
                    logger.error(f"Failed to delete {file_path}. Reason: {e}")
        logger.info(f"Successfully cleared contents of the artifact directory: {artifact_dir}")
        
    except Exception as e:
            error_message = SensorFaultException(error_message=str(e),error_detail=sys)
            logger.error(msg=f"clear_artifact_folder :: Status:Failed :: error_message:{error_message}")
            raise error_message 

def clear_dashboard_folder() -> None:
    """
    Clears the contents of the dashboard directory specified in the BaseArtifactConfig.

    This function checks if the dashboard directory exists. If it does, it removes the entire directory and 
    all of its contents using shutil.rmtree.

    Raises:
        SensorFaultException: If an error occurs while attempting to delete the dashboard directory,
        an exception is raised with details about the failure.
    """
    try:
        if os.path.exists(BaseArtifactConfig.dashboard_dir):
            shutil.rmtree(BaseArtifactConfig.dashboard_dir)
            logger.info(f"clear_dashboard_folder :: Status:Success")
        
    except Exception as e:
            error_message = SensorFaultException(error_message=str(e),error_detail=sys)
            logger.error(msg=f"clear_dashboard_folder :: Status:Failed :: error_message:{error_message}")
            raise error_message        

def create_dashboard_folder() -> None:
    """
    Creates the dashboard directory specified in the BaseArtifactConfig.

    This function attempts to create the dashboard directory. If the directory already exists,
    it will not raise an error due to the `exist_ok=True` parameter.

    Raises:
        SensorFaultException: If an error occurs while attempting to create the dashboard directory,
        an exception is raised with details about the failure.
    """
    try:
        os.makedirs(BaseArtifactConfig.dashboard_dir,exist_ok= True)
        logger.info(f"create_dashboard_folder :: Status:Success")
        
    except Exception as e:
            error_message = SensorFaultException(error_message=str(e),error_detail=sys)
            logger.error(msg=f"create_dashboard_folder :: Status:Failed :: error_message:{error_message}")
            raise error_message  
     
def remove_file(file_path:Path) -> None:
    """
    Removes a file at the specified file path.

    This function checks if the file exists at the given path, and if it does, 
    it deletes the file. If the file does not exist, no action is taken.

    Args:
        file_path (Path): The path of the file to be removed.

    Raises:
        SensorFaultException: If an error occurs while attempting to remove the file,
        an exception is raised with details about the failure.
    """
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            logger.info(f"remove_file :: Status:Deleted :: file_path : {file_path}")
        
    except Exception as e:
            error_message = SensorFaultException(error_message=str(e),error_detail=sys)
            logger.error(msg=f"clear_artifact_folder :: Status:Failed :: error_message:{error_message}")
            raise error_message
                       
def check_folder_empty(folder_path:Path) -> bool:
    """check_folder_empty :Used for check the folder path empty or not

    Args:
        folder_path (Path): Folder path for checking

    Raises:
        error_message: Custom Exception

    Returns:
        bool: status if True folder is empty else False
    """
    try:
        if os.path.exists(folder_path):
            if len(list(os.listdir(folder_path)))!=0:
                logger.info(f'check_folder_empty :: Status:Folder is not empty :: folder_path:{folder_path}')
                status = False
            else:
                logger.info(f'check_folder_empty :: Status:Folder is  empty :: folder_path:{folder_path}')
                status = True
        return status    
    except Exception as e:
            error_message = SensorFaultException(error_message=str(e),error_detail=sys)
            logger.error(msg=f"check_folder_empty :: Status:Failed :: error_message:{error_message}")
            raise error_message
                         
def save_bad_file_names() -> None:
    """
    Saves the names of bad files to a JSON file.

    This function retrieves the list of files from the specified directory 
    containing bad raw data. It then creates a folder for the JSON file (if it 
    does not already exist) and saves the list of bad file names into a JSON 
    format.

    The JSON file is stored at the path specified by 
    `PredictionRawDataValidationConfig.dashboard_bad_file_names_json_path`.

    Raises:
        SensorFaultException: If an error occurs while attempting to read the 
        file names, create a folder, or save the JSON data, an exception is raised 
        with details about the failure.
    """
    try:
        files_data = {}
        files_list = os.listdir(PredictionRawDataValidationConfig.bad_raw_data_folder_path)
        files_data['bad_files'] = files_list
        create_folder_using_file_path(PredictionRawDataValidationConfig.dashboard_bad_file_names_json_path)
        save_json(PredictionRawDataValidationConfig.dashboard_bad_file_names_json_path,files_data)
        logger.info(f"save_bad_file_names :: Status:Success")
        
    except Exception as e:
            error_message = SensorFaultException(error_message=str(e),error_detail=sys)
            logger.error(msg=f"save_bad_file_names :: Status:Failed :: error_message:{error_message}")
            raise error_message 

def get_bad_file_names() -> list[Any]:
    """
    Retrieves the names of bad files from a JSON file.

    This function reads a JSON file specified by 
    `PredictionRawDataValidationConfig.dashboard_bad_file_names_json_path`, 
    which contains a list of bad file names under the key 'bad_files'. 
    It returns this list as a Python list.

    Returns:
        list[Any]: A list containing the names of bad files.

    Raises:
        SensorFaultException: If an error occurs while reading the JSON file 
        or processing its contents, an exception is raised with details about 
        the failure.
    """
    try:
        files_data = read_json(PredictionRawDataValidationConfig.dashboard_bad_file_names_json_path)
        logger.info(f"get_bad_file_names :: Status:Success")
        return list(files_data['bad_files'])
    
    except Exception as e:
            error_message = SensorFaultException(error_message=str(e),error_detail=sys)
            logger.error(msg=f"get_bad_file_names :: Status:Failed :: error_message:{error_message}")
            raise error_message
                             
def get_local_file_md5(file_path) -> str:
    """
    Calculates the MD5 hash of a local file.

    This function opens a file located at the specified `file_path`, reads it in chunks, 
    and computes its MD5 hash. The resulting hash is returned as a hexadecimal string.

    Args:
        file_path (str): The path to the local file for which the MD5 hash is to be computed.

    Returns:
        str: The MD5 hash of the file, represented as a hexadecimal string.

    Raises:
        SensorFaultException: If an error occurs while opening the file or computing the hash, 
        an exception is raised with details about the failure.
    """
    try:
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        logger.info(f"get_local_file_md5 :: Status:Success :: HastTag:{hash_md5.hexdigest()}")        
        return hash_md5.hexdigest()    
    
    except Exception as e:
            error_message = SensorFaultException(error_message=str(e),error_detail=sys)
            logger.error(msg=f"get_local_file_md5 :: Status:Failed :: error_message:{error_message}")
            raise error_message

def save_confusion_matrix_image(y_test,y_pred,image_path:Path) -> None:
    """
    Generates and saves a confusion matrix image for the given true and predicted labels.

    This function computes the confusion matrix based on the true labels (`y_test`) 
    and the predicted labels (`y_pred`). It visualizes the confusion matrix using Matplotlib 
    and saves the plot to the specified image path.

    Args:
        y_test (array-like): True labels of the test dataset.
        y_pred (array-like): Predicted labels generated by the model.
        image_path (Path): The path where the confusion matrix image will be saved.

    Raises:
        SensorFaultException: If an error occurs during the computation of the confusion matrix,
        visualization, or saving the image, an exception is raised with details about the failure.
    """
    try:
        # Generate confusion matrix
        cm = confusion_matrix(y_test, y_pred)

        # Display confusion matrix
        disp = ConfusionMatrixDisplay(confusion_matrix=cm, display_labels=['Class 0', 'Class 1'])
        plt.figure(figsize=(8, 6))
        disp.plot(cmap='Blues', values_format='d')  # Use values_format='d' for integers
        plt.title('Confusion Matrix')

        # Save the plot as an image
        plt.savefig(image_path, bbox_inches='tight')
        plt.close()    
    
    except Exception as e:
            error_message = SensorFaultException(error_message=str(e),error_detail=sys)
            logger.error(msg=f"get_local_file_md5 :: Status:Failed :: error_message:{error_message}")
            raise error_message        
                          
def send_datadrift_mail_team(url:str,s3_prediction_folder_name:str) -> None:
    """
    Sends an email notification to the data science team regarding detected data drift.

    This function constructs an HTML email alerting the team about significant data drift 
    identified in model performance. The email includes a link to review the full data drift 
    report and mentions the associated S3 prediction folder containing relevant data files.

    Args:
        url (str): The URL link to the data drift report for review.
        s3_prediction_folder_name (str): The name of the S3 prediction folder containing 
                                           additional CSV files and data.

    Raises:
        SensorFaultException: If an error occurs during the email composition, connection to 
        the SMTP server, or sending the email, an exception is raised with details about the 
        failure.

    Returns:
        None
    """
    try:
        # Email configuration
        sender_email = os.getenv('SENDER_EMAIL')
        sender_password = os.getenv('SENDER_EMAIL_PASSWORD')         
        recipient_email = os.getenv('RECIPIENT_EMAIL')
        
        subject = "Data Drift Report"
    
        # body of email
        html_body = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                @keyframes blink {{
                    0% {{ opacity: 1; }}
                    50% {{ opacity: 0.5; }}
                    100% {{ opacity: 1; }}
                }}
                body {{
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    background-color: #f0f4f8;
                    padding: 20px;
                }}
                .container {{
                    background-color: #ffffff;
                    border-radius: 10px;
                    padding: 30px;
                    box-shadow: 0 4px 10px rgba(0, 0, 0, 0.1);
                    max-width: 600px;
                    margin: 0 auto;
                }}
                h2 {{
                    font-size: 24px;
                    color: #2c3e50;
                    text-align: center;
                }}
                .alert {{
                    background-color: #fff3cd;
                    border: 1px solid #ffcc00;
                    padding: 20px;
                    border-radius: 8px;
                    font-size: 18px;
                    font-weight: bold;
                    color: #d9534f;
                    text-align: center;
                    animation: blink 1.5s infinite;
                }}
                .alert-title {{
                    font-size: 20px;
                    color: #d9534f;
                    font-weight: bold;
                }}
                p {{
                    font-size: 16px;
                    color: #2c3e50;
                    line-height: 1.6;
                }}
                .footer {{
                    margin-top: 30px;
                    font-size: 14px;
                    color: #999;
                    text-align: center;
                }}
                .button {{
                    background-color: #007bff;
                    color: white;
                    padding: 12px 24px;
                    border-radius: 5px;
                    text-decoration: none;
                    font-weight: bold;
                    display: inline-block;
                    margin: 20px 0;
                    text-align: center;
                }}
                .button:hover {{
                    background-color: #0056b3;
                }}
                .subject {{
                    color: red;
                    font-weight: bold;
                    text-align: center;
                    font-size: 22px;
                    margin: 20px 0;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <h2>🚨 Data Drift Alert</h2>
                <div class="subject">Critical: Data Drift Detected</div>
                <div class="alert">
                    <div class="alert-title">Action Required</div>
                    <p>Our latest analysis has identified significant data drift that may impact model performance. Immediate review is recommended to prevent potential inaccuracies in future predictions.</p>
                </div>
                <p>To access the full data drift report and take appropriate measures, please use the link below:</p>
                <a href="{url}" class="button">Review Data Drift Report</a>

                <p>For your reference, please check the retraining folder and the prediction folder <strong>{s3_prediction_folder_name}</strong> in your S3 bucket for additional CSV files and data.</p>
                
                <div class="footer">
                    <p>Thank you for your attention.<br>Best regards,<br>Your Data Science Team</p>
                </div>
            </div>
        </body>
        </html>
        """
        # Create the email message
        msg = MIMEMultipart()
        msg['From'] = sender_email # type: ignore
        msg['To'] = recipient_email # type: ignore
        msg['Subject'] = subject

        # Attach the HTML body
        msg.attach(MIMEText(html_body, 'html'))

        # Send the email
        try:
            with smtplib.SMTP('smtp.gmail.com', 587, timeout=1500) as server:
                server.starttls()  # Upgrade the connection to a secure encrypted SSL/TLS
                server.login(sender_email, sender_password)  # Log in to the email server # type: ignore
                server.send_message(msg)  # Send the email
            logger.info("Data Drift Email sent successfully!")
        except Exception as e:
            logger.info(f"Failed to send Data Drift Email: {e}")                             
        
        logger.info(msg=f"send_datadrift_mail to backend team :: Status:Success")
        
    except Exception as e:
            error_message = SensorFaultException(error_message=str(e),error_detail=sys)
            logger.error(msg=f"send_datadrift_mail :: Status:Failed :: error_message:{error_message}")
            raise error_message         
        
def send_datadrift_mail_client(excel_file_path: str) -> None:
    """
    Sends an email to the client notifying them of detected data drift and requesting feedback.

    This function constructs an HTML email containing information about the data drift detected 
    in the model predictions, along with an attached Excel file containing the relevant data.
    The email is sent using the SMTP protocol through a Gmail account.

    Parameters:
    ----------
    excel_file_path : str
        The file path to the Excel file that contains the predictions or data requiring client feedback.

    Returns:
    -------
    None

    Raises:
    ------
    SensorFaultException
        If an error occurs during the process of sending the email, including issues with 
        environment variables, file handling, or SMTP operations.

    Notes:
    ------
    - Ensure that the environment variables for SENDER_EMAIL, SENDER_EMAIL_PASSWORD, and 
      RECIPIENT_EMAIL are set correctly before calling this function.
    - The email includes an alert indicating the urgency of the feedback required for data 
      drift management.
    """
    try:
        # Email configuration
        sender_email = os.getenv('SENDER_EMAIL')
        sender_password = os.getenv('SENDER_EMAIL_PASSWORD')         
        recipient_email = os.getenv('RECIPIENT_EMAIL')
        
        subject = "Urgent: Data Drift Detected - Client Feedback Required"

        # Body of the email
        html_body = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                @keyframes fadeIn {{
                    0% {{ opacity: 0; }}
                    100% {{ opacity: 1; }}
                }}
                @keyframes bounce {{
                    0%, 20%, 50%, 80%, 100% {{ transform: translateY(0); }}
                    40% {{ transform: translateY(-20px); }}
                    60% {{ transform: translateY(-10px); }}
                }}
                body {{
                    font-family: 'Arial', sans-serif;
                    background-color: #f4f7f9;
                    padding: 20px;
                }}
                .container {{
                    background-color: #ffffff;
                    border-radius: 12px;
                    padding: 40px;
                    box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
                    max-width: 600px;
                    margin: 0 auto;
                    animation: fadeIn 2s;
                }}
                h2 {{
                    font-size: 26px;
                    color: #2c3e50;
                    text-align: center;
                    margin-bottom: 20px;
                }}
                .alert {{
                    background-color: #ffdfdf;
                    border-left: 4px solid #ff6f6f;
                    padding: 20px;
                    border-radius: 8px;
                    font-size: 18px;
                    color: #d9534f;
                    text-align: center;
                    animation: bounce 2s infinite;
                }}
                .alert-title {{
                    font-size: 22px;
                    color: #d9534f;
                    font-weight: bold;
                }}
                p {{
                    font-size: 16px;
                    color: #2c3e50;
                    line-height: 1.6;
                }}
                .footer {{
                    margin-top: 30px;
                    font-size: 14px;
                    color: #999;
                    text-align: center;
                }}
                .button {{
                    background-color: #28a745;
                    color: white;
                    padding: 12px 24px;
                    border-radius: 5px;
                    text-decoration: none;
                    font-weight: bold;
                    display: inline-block;
                    margin: 20px 0;
                    text-align: center;
                }}
                .button:hover {{
                    background-color: #218838;
                }}
                .subject {{
                    color: red;
                    font-weight: bold;
                    text-align: center;
                    font-size: 22px;
                    margin: 20px 0;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <h2>🚨 Data Drift Detected - Feedback Required</h2>
                <div class="alert">
                    <div class="alert-title">Action Required: Provide Feedback</div>
                    <p>We have detected a significant change in the data distribution. Please review the attached prediction file and provide feedback on the new data.<br><strong>Note:</strong> Confidence Score below 0.75 is mandatory</p>
                </div>
                <p>Your feedback will be essential in retraining the model to ensure its accuracy and reliability for future predictions.</p>
                <p>We appreciate your timely response in this matter. Thank you for your cooperation.</p>
                <div class="footer">
                    <p>Best regards,<br>(Tech Team)</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Create the email message
        msg = MIMEMultipart()
        msg['From'] = sender_email # type: ignore
        msg['To'] = recipient_email # type: ignore
        msg['Subject'] = subject

        # Attach the HTML body
        msg.attach(MIMEText(html_body, 'html'))

        # Attach the Excel file
        with open(excel_file_path, "rb") as file:
            part = MIMEApplication(file.read(), Name=os.path.basename(excel_file_path))
            part['Content-Disposition'] = f'attachment; filename="{os.path.basename(excel_file_path)}"'
            msg.attach(part)

        # Send the email
        try:
            with smtplib.SMTP('smtp.gmail.com', 587) as server:
                server.starttls()  # Upgrade the connection to a secure encrypted SSL/TLS
                server.login(sender_email, sender_password)  # Log in to the email server # type: ignore
                server.send_message(msg)  # Send the email
            logger.info("Data Drift Email with attached file sent successfully!")
        except Exception as e:
            logger.info(f"Failed to send Data Drift Email: {e}")

        logger.info(msg=f"send_datadrift_mail to client :: Status:Success")

    except Exception as e:
        error_message = SensorFaultException(error_message=str(e), error_detail=sys)
        logger.error(msg=f"send_datadrift_mail :: Status:Failed :: error_message:{error_message}")
        raise error_message    